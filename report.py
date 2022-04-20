import os
import sys

from PyQt5.QtWidgets import QApplication, QWidget, QFileDialog, QMessageBox
from ui_report import Ui_Dialog

from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Border, Font
import aspose.words as aw


class ReportWindow(QWidget):
    def __init__(self):
        super(ReportWindow, self).__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.dir = ''

        self.ui.btn_chouse_dir.clicked.connect(self.get_dir)
        self.ui.btn_report.clicked.connect(self.report)
        self.ui.btn_cancel.clicked.connect(self.close)

        self.show()

    def get_dir(self):
        selected_dir = QFileDialog.getExistingDirectory(self, caption='Choose Directory', directory=os.getcwd())
        if selected_dir:
            self.dir = selected_dir
            self.ui.le_folder_path.setText(selected_dir)

    def report(self):
        if not self.dir:
            self.showDialogError("Ошибка: Выберите путь.")
        elif not (self.ui.checkBox_word.isChecked() or self.ui.checkBox_excel.isChecked()):
            self.showDialogError("Ошибка: Выберите нужный формат файла.")
        else:
            data = []
            try:
                Fin = open('case.txt', 'r')
                for line in Fin.readlines():
                    line = line.split()
                    data.append([line[0], ' '.join(map(str, line[1:-2])), line[-1]])
            except IOError:
                sys.exit()
            finally:
                Fin.close()

            if self.ui.checkBox_excel.isChecked():
                self.create_excel(data, self.dir)
            elif self.ui.checkBox_word.isChecked():
                self.create_word()
            elif self.ui.checkBox_word.isChecked() and self.ui.checkBox_excel.isChecked():
                self.create_excel(data, self.dir)
                self.create_word()
            self.showDialogError("Отчет создан")
            self.close()

    def create_excel(self, data, path_file):
        # Создаем новый лист
        workbook = Workbook()
        ws = workbook.active

        # Стили шрифта
        font_title = Font(name='Times New Roman', size=14, bold=True)
        font_table = Font(name='Times New Roman', size=12)

        # Заголовок таблицы
        ws['A1'] = '№\nп/п'
        ws['B1'] = 'Наименование показателя'
        ws['C1'] = 'Значение показателя'

        # Горизонтальное и вертикальное выравнивание текста
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Оформление границ ячеек
        ws.row_dimensions[1].height = 33
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15

        ws.row_dimensions[3].height = 30
        ws.row_dimensions[9].height = 30
        ws.row_dimensions[10].height = 30

        for i in range(4, 9):
            ws.row_dimensions[i].height = 40

        # Стиль текста ячеек
        ws['A1'].font = font_title
        ws['B1'].font = font_title
        ws['C1'].font = font_title

        # Заполняем таблицу
        row = 2

        for inum, text, value in data:
            ws[f'A{row}'] = inum
            ws[f'A{row}'].font = font_table
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')

            ws[f'B{row}'] = text
            ws[f'B{row}'].font = font_table
            ws[f'B{row}'].alignment = Alignment(
                horizontal='general',
                vertical='bottom',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0
            )

            ws[f'C{row}'] = value
            ws[f'C{row}'].font = font_table
            ws[f'C{row}'].alignment = Alignment(horizontal='right')

            row += 1

        # Устанавливаем границы таблицы
        thin = Side(border_style="thin", color="000000")
        for row in ws[f'A1:C{len(data) + 1}']:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Сохраняем файл
        workbook.save(f"{path_file}/Образовательная деятельность СПО.xlsx")

    def create_word(self):
        pass

    def showDialogError(self, text):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(text)
        msgBox.setWindowTitle("Ошибка")
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec()


def create_word():
    # Create a new Word document.
    doc = aw.Document()

    # Create document builder.
    builder = aw.DocumentBuilder(doc)

    # Start the table.
    table = builder.start_table()

    # Insert cell.
    builder.insert_cell()

    # Table wide formatting must be applied after at least one row is present in the table.
    table.left_indent = 20.0

    # Set height and define the height rule for the header row.
    builder.row_format.height = 40.0
    builder.row_format.height_rule = aw.HeightRule.AT_LEAST

    # Set alignment and font settings.
    builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
    builder.font.size = 16
    builder.font.name = "Arial"
    builder.font.bold = True

    builder.cell_format.width = 100.0
    builder.write("Header Row,\n Cell 1")

    # We don't need to specify this cell's width because it's inherited from the previous cell.
    builder.insert_cell()
    builder.write("Header Row,\n Cell 2")

    builder.insert_cell()
    builder.cell_format.width = 200.0
    builder.write("Header Row,\n Cell 3")
    builder.end_row()

    builder.cell_format.width = 100.0
    builder.cell_format.vertical_alignment = aw.tables.CellVerticalAlignment.CENTER

    # Reset height and define a different height rule for table body.
    builder.row_format.height = 30.0
    builder.row_format.height_rule = aw.HeightRule.AUTO
    builder.insert_cell()

    # Reset font formatting.
    builder.font.size = 12
    builder.font.bold = False

    builder.write("Row 1, Cell 1 Content")
    builder.insert_cell()
    builder.write("Row 1, Cell 2 Content")

    builder.insert_cell()
    builder.cell_format.width = 200.0
    builder.write("Row 1, Cell 3 Content")
    builder.end_row()

    builder.insert_cell()
    builder.cell_format.width = 100.0
    builder.write("Row 2, Cell 1 Content")

    builder.insert_cell()
    builder.write("Row 2, Cell 2 Content")

    builder.insert_cell()
    builder.cell_format.width = 200.0
    builder.write("Row 2, Cell 3 Content.")
    builder.end_row()

    # End table.
    builder.end_table()

    # Save the document.
    doc.save('/home/yaroslove/test/Word.docx')


def main():
    app = QApplication(sys.argv)
    ex = ReportWindow()
    sys.exit(app.exec_())


if __name__ == '__main__':
    create_word()
